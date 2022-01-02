print("Please wait... \n This might take a while")

import openpyxl as xl

#importing Data frome Excel
wb=xl.load_workbook("Perodic Table Elements Data.xlsx")
datasheet=wb['Elements']
rows=datasheet.max_row

#Making a empty Dictonary to store all input data
elements={}

#loading all excel data to dictionary with all possible input (Atomicnumber , Element name , Symbol)
for i in range(2,rows+2):
    #atomoc number as input
    elements[datasheet.cell(i,1).value]=[datasheet.cell(i,1).value,
                                         datasheet.cell(i,2).value,
                                         datasheet.cell(i,3).value,
                                         datasheet.cell(i,4).value,
                                         datasheet.cell(i,5).value,
                                         datasheet.cell(i,6).value,
                                         datasheet.cell(i,7).value,
                                         datasheet.cell(i,8).value,
                                         datasheet.cell(i,9).value,
                                         datasheet.cell(i,10).value,
                                         datasheet.cell(i,11).value,
                                         datasheet.cell(i,12).value,
                                         datasheet.cell(i,13).value,
                                         datasheet.cell(i,14).value,
                                         datasheet.cell(i,15).value,
                                         datasheet.cell(i,16).value,
                                         datasheet.cell(i,17).value,
                                         datasheet.cell(i,18).value,
                                         datasheet.cell(i,19).value,
                                         datasheet.cell(i,20).value,
                                         datasheet.cell(i,21).value,
                                         datasheet.cell(i,22).value,
                                         datasheet.cell(i,23).value]

for i in range(2, rows + 2):
#element name as input
    elements[datasheet.cell(i, 2).value] = [datasheet.cell(i, 1).value,
                                            datasheet.cell(i, 2).value,
                                            datasheet.cell(i, 3).value,
                                            datasheet.cell(i, 4).value,
                                            datasheet.cell(i, 5).value,
                                            datasheet.cell(i, 6).value,
                                            datasheet.cell(i, 7).value,
                                            datasheet.cell(i, 8).value,
                                            datasheet.cell(i, 9).value,
                                            datasheet.cell(i, 10).value,
                                            datasheet.cell(i, 11).value,
                                            datasheet.cell(i, 12).value,
                                            datasheet.cell(i, 13).value,
                                            datasheet.cell(i, 14).value,
                                            datasheet.cell(i, 15).value,
                                            datasheet.cell(i, 16).value,
                                            datasheet.cell(i, 17).value,
                                            datasheet.cell(i, 18).value,
                                            datasheet.cell(i, 19).value,
                                            datasheet.cell(i, 20).value,
                                            datasheet.cell(i, 21).value,
                                            datasheet.cell(i, 22).value,
                                            datasheet.cell(i, 23).value]
for i in range(2, rows + 2):
#Symbol as input
    elements[datasheet.cell(i, 3).value] = [datasheet.cell(i, 1).value,
                                            datasheet.cell(i, 2).value,
                                            datasheet.cell(i, 3).value,
                                            datasheet.cell(i, 4).value,
                                            datasheet.cell(i, 5).value,
                                            datasheet.cell(i, 6).value,
                                            datasheet.cell(i, 7).value,
                                            datasheet.cell(i, 8).value,
                                            datasheet.cell(i, 9).value,
                                            datasheet.cell(i, 10).value,
                                            datasheet.cell(i, 11).value,
                                            datasheet.cell(i, 12).value,
                                            datasheet.cell(i, 13).value,
                                            datasheet.cell(i, 14).value,
                                            datasheet.cell(i, 15).value,
                                            datasheet.cell(i, 16).value,
                                            datasheet.cell(i, 17).value,
                                            datasheet.cell(i, 18).value,
                                            datasheet.cell(i, 19).value,
                                            datasheet.cell(i, 20).value,
                                            datasheet.cell(i, 21).value,
                                            datasheet.cell(i, 22).value,
                                            datasheet.cell(i, 23).value]

#While loop so that the programm will run until user says no
while True:
    # caution for user
    print("While you enter your input make sure you Enter the first letter as capital \n")

    # taking input
    User_input = input("Enter an Element name or Symbol or Atomic Number of an element or quit to exit programm:")

    # trying to convert user input to integer if the atomic number if input is Atomic number
    try:
        User_input = int(User_input)
    except:
        EOFError


    # Checking if the input is correct

    # if input is found in datasheet
    if User_input in elements:
        # making a list of the values of the input

        elementsdata = elements[User_input]

        print("Data Found \n")
        print("Elemement name :", elementsdata[1])
        print("Atomic Number :", elementsdata[0])
        print("Symbol :", elementsdata[2])
        print("Atomic Mass :", elementsdata[3])
        print("--------------")
        print("Number of Neutrons :", elementsdata[4])
        print("Number of Protons :", elementsdata[5])
        print("Number of Electrons :", elementsdata[6])
        print("--------------")
        print("Period :", elementsdata[7])
        print("Group :", elementsdata[8])
        print("--------------")
        print("State of Matter :", elementsdata[9])
        print("--------------")
        if elementsdata[10] == "yes":
            print("Radioactive : yes")
        else:
            print("Radioactive : no")
        if elementsdata[11] == "yes":
            print("Natural : Yes")
        else:
            print("Natural : no")
        print("--------------")
        if elementsdata[12] == "yes":
            print("Metal / Non-metal / Metalloid : Metal")
        elif elementsdata[13] == "yes":
            print("Metal / Non-metal / Metalloid : Non-Metal")
        elif elementsdata[12] == "yes":
            print("Metal / Non-metal / Metalloid : Metealloid")
        else:
            print("Metal / Non-metal / Metalloid : ---")
        print("Type :", elementsdata[15])
        print("--------------")
        print("Atomic Radius :", elementsdata[16])
        print("--------------")
        print("Melting Point :", elementsdata[17])
        print("Boiling Point :", elementsdata[18])
        print("--------------")
        print("Discoverd by ", elementsdata[19], " in year ", elementsdata[20])
        print("--------------")
        print("Number of Shells :", elementsdata[21])
        print("Number if Valancie Electrons", elementsdata[22])
    # quit program
    elif User_input == "quit" or User_input == "Quit":
        break
    # if input is not found in datasheet
    else:
        print(
            "\n Couldnt find the input in datasheet please check spelling or caps \n even when you enter Symbol or element name make sure first letter is capital")

#ending and exiting programm
exit("\t Programm ended \n \t   Thankyou!")
